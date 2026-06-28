import base64
import json
import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# =========================================================
# TOPAZ SMART DOCUMENT TRACKER V6.1 ENTERPRISE
# Shared dashboard:
# - Admin can upload and generate.
# - Viewer can see the latest generated dashboard in read-only mode.
# - Added: Status Summary by Category from Sheet RFA / RFI.
# =========================================================

st.set_page_config(
    page_title="Topaz Smart Document Tracker",
    page_icon="💎",
    layout="wide",
)

TRACKING_SHEETS = ["RFA", "RFI"]
TAKENAKA_SHEETS = ["MAT_ICT", "DWG_ICT", "MTS_ICT"]

LOGO_PATH = "assets/logo.png"

DATA_DIR = Path("data")
LATEST_CSV = DATA_DIR / "latest_result.csv"
LATEST_META = DATA_DIR / "latest_meta.json"
LATEST_REPORT = DATA_DIR / "latest_report.xlsx"
LATEST_STATUS_SUMMARY = DATA_DIR / "latest_status_summary.csv"
LATEST_STATUS_DETAIL = DATA_DIR / "latest_status_detail.csv"

DATA_DIR.mkdir(parents=True, exist_ok=True)


# =========================================================
# STYLE
# =========================================================
st.markdown(
    """
<style>
.stApp {
    background: #f4f7fb;
    color: #0f172a;
}

.block-container {
    max-width: 1500px;
    padding-top: 1.2rem;
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #061124 0%, #0f172a 70%, #111827 100%);
}

[data-testid="stSidebar"] * {
    color: white;
}

.sidebar-logo-title {
    font-size: 31px;
    font-weight: 900;
    margin-top: 8px;
    letter-spacing: .5px;
}

.sidebar-subtitle {
    color: #c7d2fe;
    font-size: 13px;
    margin-bottom: 20px;
}

.sidebar-card {
    padding: 16px;
    border-radius: 18px;
    background: rgba(255,255,255,.08);
    border: 1px solid rgba(255,255,255,.14);
    margin: 16px 0;
}

.nav-active {
    padding: 11px 13px;
    border-radius: 13px;
    margin: 8px 0;
    background: linear-gradient(90deg,#4f46e5,#7c3aed);
    font-weight: 800;
}

.nav-item {
    padding: 11px 13px;
    border-radius: 13px;
    margin: 8px 0;
    background: rgba(255,255,255,.06);
    font-weight: 700;
}

.hero {
    background: linear-gradient(90deg, #071226 0%, #0b1b4d 50%, #172554 100%);
    border-radius: 26px;
    padding: 30px 34px;
    color: white;
    box-shadow: 0 20px 45px rgba(15, 23, 42, .22);
    margin-bottom: 22px;
}

.hero-grid {
    display: grid;
    grid-template-columns: auto 1fr auto;
    gap: 22px;
    align-items: center;
}

.hero-logo {
    width: 92px;
    height: 92px;
    object-fit: contain;
}

.hero-title {
    font-size: 40px;
    font-weight: 950;
    line-height: 1.05;
    margin-bottom: 8px;
}

.hero-subtitle {
    color: #dbeafe;
    font-size: 15px;
}

.hero-badge {
    padding: 12px 18px;
    border-radius: 999px;
    background: rgba(255,255,255,.13);
    border: 1px solid rgba(255,255,255,.22);
    font-weight: 800;
    white-space: nowrap;
}

.info-box {
    padding: 16px 18px;
    border-radius: 18px;
    background: linear-gradient(90deg, #ecfdf5, #ffffff);
    border: 1px solid #bbf7d0;
    color: #166534;
    font-weight: 800;
    margin: 18px 0;
}

.viewer-box {
    padding: 16px 18px;
    border-radius: 18px;
    background: linear-gradient(90deg, #eff6ff, #ffffff);
    border: 1px solid #bfdbfe;
    color: #1d4ed8;
    font-weight: 800;
    margin: 18px 0;
}

.warning-box {
    padding: 16px 18px;
    border-radius: 18px;
    background: linear-gradient(90deg, #fff7ed, #ffffff);
    border: 1px solid #fed7aa;
    color: #9a3412;
    font-weight: 800;
    margin: 18px 0;
}

.kpi-card {
    padding: 24px;
    border-radius: 22px;
    background: white;
    box-shadow: 0 14px 32px rgba(15, 23, 42, 0.08);
    border: 1px solid #e5e7eb;
    min-height: 150px;
    position: relative;
    overflow: hidden;
}

.kpi-card:after {
    content:"";
    position:absolute;
    left:0;
    right:0;
    bottom:0;
    height:5px;
    background: var(--accent);
}

.kpi-icon {
    font-size: 26px;
    margin-bottom: 14px;
}

.kpi-title {
    font-size: 13px;
    color: #64748b;
    font-weight: 850;
}

.kpi-value {
    font-size: 38px;
    font-weight: 950;
    color: #0f172a;
    line-height: 1.1;
    margin: 7px 0;
}

.kpi-sub {
    color: #64748b;
    font-size: 13px;
}

.panel {
    background: white;
    border:1px solid #e2e8f0;
    border-radius: 24px;
    box-shadow: 0 14px 32px rgba(15,23,42,.07);
    padding: 24px;
    margin-bottom: 18px;
}

.panel-title {
    font-size: 22px;
    font-weight: 950;
    color:#0f172a;
    margin-bottom: 16px;
}

.quick-row {
    display:flex;
    justify-content:space-between;
    align-items:center;
    padding:13px 14px;
    border-radius:14px;
    background:#f8fafc;
    border:1px solid #e2e8f0;
    margin-bottom:10px;
}

.count-pill {
    padding:4px 11px;
    border-radius:999px;
    font-weight:900;
    background:#dbeafe;
    color:#1d4ed8;
}

[data-testid="stFileUploaderDropzone"] {
    border-radius: 18px;
    border: 1px solid #dbe3ef;
    background: #f8fafc;
}

.stButton > button {
    border-radius: 14px;
    font-weight: 800;
}
</style>
""",
    unsafe_allow_html=True,
)


# =========================================================
# UI HELPERS
# =========================================================
def image_to_base64(path: str) -> str:
    if not os.path.exists(path):
        return ""

    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def render_logo_html(size_class: str = "hero-logo") -> str:
    logo64 = image_to_base64(LOGO_PATH)

    if logo64:
        return f'<img class="{size_class}" src="data:image/png;base64,{logo64}">'

    return '<div style="font-size:62px;">💎</div>'


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Summary") -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    output.seek(0)
    return output.getvalue()


# =========================================================
# NEW: STATUS SUMMARY FROM RFA / RFI
# =========================================================
def normalize_status(value):
    text = str(value or "").strip()

    status_map = {
        "open": "Open",
        "on progress": "On Progress",
        "on process": "On Progress",
        "in progress": "On Progress",
        "approved": "Approved",
        "approve": "Approved",
        "close": "Close",
        "closed": "Close",
    }

    return status_map.get(text.lower(), text)


def get_column_by_name_or_position(ws, preferred_name, fallback_letter):
    """
    Find column by header name in row 1.
    If not found, use fallback Excel column letter.
    """
    preferred_name = preferred_name.strip().lower()

    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == preferred_name:
            return cell.column_letter

    return fallback_letter


def extract_revision_value(value):
    """
    Convert revision/version value to comparable number.
    Examples:
    00 -> 0
    01 -> 1
    Rev02 -> 2
    A -> 1001
    B -> 1002
    Blank -> 0
    """
    text = str(value or "").strip()

    if not text or text.lower() == "nan":
        return 0

    numbers = re.findall(r"\d+", text)
    if numbers:
        return int(numbers[-1])

    if len(text) == 1 and text.isalpha():
        return 1000 + ord(text.upper()) - ord("A") + 1

    return 0


def build_status_summary_from_tracking(tracking_file):
    """
    V6.1.5 Logic:
    1) Read Sheet RFA and RFI.
    2) Use Document No + Category + Status + Revision.
    3) Keep only latest revision per Document No.
    4) Count Status by Category.
    """
    wb = load_workbook(tracking_file, data_only=True)

    records = []

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        category_col = get_column_by_name_or_position(ws, "Category", "C")
        status_col = get_column_by_name_or_position(ws, "Status", "F")
        revision_col = get_column_by_name_or_position(ws, "Rev", "E")

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"B{row}"].value
            category = ws[f"{category_col}{row}"].value
            doc_name = ws[f"D{row}"].value
            revision = ws[f"{revision_col}{row}"].value
            status = ws[f"{status_col}{row}"].value
            info = ws[f"G{row}"].value

            doc_no_text = str(doc_no or "").strip()
            category_text = str(category or "").strip()
            status_text = normalize_status(status)

            if not doc_no_text or doc_no_text.lower() == "nan":
                continue

            if not category_text or category_text.lower() == "nan":
                continue

            if not status_text or status_text.lower() == "nan":
                continue

            records.append(
                {
                    "Tracking Sheet": sheet,
                    "Document No": doc_no_text,
                    "Category": category_text,
                    "Document Name": doc_name,
                    "Revision": str(revision or "").strip(),
                    "Revision Sort": extract_revision_value(revision),
                    "Status": status_text,
                    "Info": info,
                    "Excel Row": row,
                }
            )

    detail_all_df = pd.DataFrame(records)

    if detail_all_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Keep only latest revision per Document No
    # If revision is the same, keep the lower/latest Excel row order by sorting Excel Row as well.
    latest_df = (
        detail_all_df
        .sort_values(["Document No", "Revision Sort", "Excel Row"])
        .groupby("Document No", as_index=False)
        .tail(1)
        .copy()
    )

    keep_status = ["Open", "On Progress", "Approved"]
    df = latest_df[latest_df["Status"].isin(keep_status)].copy()

    categories = sorted(df["Category"].dropna().unique().tolist())

    rows = []

    for status in keep_status:
        status_df = df[df["Status"] == status]

        row = {"Status": status, "Total": len(status_df)}
        for cat in categories:
            row[cat] = len(status_df[status_df["Category"] == cat])
        rows.append(row)

    total_row = {"Status": "Total Document", "Total": len(df)}
    for cat in categories:
        total_row[cat] = len(df[df["Category"] == cat])
    rows.append(total_row)

    summary_df = pd.DataFrame(rows)

    preferred_cols = ["Status", "Total", "MAT", "MCR", "MTS", "CVI", "DWG"]
    final_cols = [c for c in preferred_cols if c in summary_df.columns]
    other_cols = [c for c in summary_df.columns if c not in final_cols]
    summary_df = summary_df[final_cols + other_cols]

    # Make dashboard cleaner: show 0 as "-"
    display_df = summary_df.copy()
    for col in display_df.columns:
        if col != "Status":
            display_df[col] = display_df[col].apply(lambda x: "-" if int(x) == 0 else int(x))

    # Detail export/view should also use latest revision only
    latest_detail_df = latest_df.drop(columns=["Revision Sort", "Excel Row"], errors="ignore")

    return display_df, latest_detail_df


# =========================================================
# SHARED DATA STORAGE
# =========================================================
def save_latest_dashboard(
    df: pd.DataFrame,
    report: BytesIO,
    total_docs: int,
    open_docs: int,
    action_counts: dict,
    status_summary_df: pd.DataFrame,
    status_detail_df: pd.DataFrame,
) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    df.to_csv(LATEST_CSV, index=False, encoding="utf-8-sig")

    if status_summary_df is not None and not status_summary_df.empty:
        status_summary_df.to_csv(LATEST_STATUS_SUMMARY, index=False, encoding="utf-8-sig")

    if status_detail_df is not None and not status_detail_df.empty:
        status_detail_df.to_csv(LATEST_STATUS_DETAIL, index=False, encoding="utf-8-sig")

    with open(LATEST_REPORT, "wb") as f:
        f.write(report.getvalue())

    meta = {
        "total_docs": int(total_docs),
        "open_docs": int(open_docs),
        "action_counts": action_counts,
        "last_updated": datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
    }

    with open(LATEST_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def load_latest_dashboard():
    if not LATEST_CSV.exists() or not LATEST_META.exists():
        return None, None, None, None, None

    df = pd.read_csv(LATEST_CSV)

    with open(LATEST_META, "r", encoding="utf-8") as f:
        meta = json.load(f)

    report_bytes = LATEST_REPORT.read_bytes() if LATEST_REPORT.exists() else None

    status_summary_df = None
    if LATEST_STATUS_SUMMARY.exists():
        status_summary_df = pd.read_csv(LATEST_STATUS_SUMMARY)

    status_detail_df = None
    if LATEST_STATUS_DETAIL.exists():
        status_detail_df = pd.read_csv(LATEST_STATUS_DETAIL)

    return df, meta, report_bytes, status_summary_df, status_detail_df


def hydrate_session_from_latest() -> None:
    if st.session_state.result_df is not None:
        return

    df, meta, report_bytes, status_summary_df, status_detail_df = load_latest_dashboard()

    if df is None or meta is None:
        return

    st.session_state.result_df = df
    st.session_state.report = report_bytes
    st.session_state.total_docs = int(meta.get("total_docs", 0))
    st.session_state.open_docs = int(meta.get("open_docs", 0))
    st.session_state.action_counts = meta.get("action_counts", {})
    st.session_state.last_updated = meta.get("last_updated", "")
    st.session_state.status_summary_df = status_summary_df
    st.session_state.status_detail_df = status_detail_df


# =========================================================
# LOGIN
# =========================================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "role" not in st.session_state:
    st.session_state.role = "viewer"

if "username" not in st.session_state:
    st.session_state.username = ""


with st.sidebar:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=150)
    else:
        st.markdown("## 💎")

    st.markdown('<div class="sidebar-logo-title">TOPAZ</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-subtitle">Smart Document Tracker V6.1.5.5.4</div>', unsafe_allow_html=True)

    st.divider()

    if not st.session_state.logged_in:
        username = st.text_input("Username")

        if st.button("Login", use_container_width=True):
            st.session_state.username = username.strip()

            if username.strip().lower() in ["pavinee", "admin"]:
                st.session_state.role = "admin"
            else:
                st.session_state.role = "viewer"

            st.session_state.logged_in = True
            st.rerun()

    else:
        st.markdown(
            f"""
            <div class="sidebar-card">
                <b>🔑 Role</b><br>{st.session_state.role.title()}
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("Logout", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.role = "viewer"
            st.session_state.username = ""
            st.rerun()

    st.divider()
    st.markdown('<div class="nav-active">🏠 Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="nav-item">📋 Documents</div>', unsafe_allow_html=True)
    st.markdown('<div class="nav-item">📊 Action Summary</div>', unsafe_allow_html=True)
    st.markdown('<div class="nav-item">⬇ Download Report</div>', unsafe_allow_html=True)
    st.divider()
    st.caption("Shared Dashboard")


# =========================================================
# ORIGINAL WORKING LOGIC
# =========================================================
def base_doc_no(doc_no):
    doc_no = str(doc_no).strip()
    parts = doc_no.split("-")

    if parts[-1].isdigit() and len(parts[-1]) == 2:
        return "-".join(parts[:-1])

    return doc_no


def read_takenaka(takenaka_file):
    wb = load_workbook(takenaka_file, data_only=True)
    data = {}

    for sheet in TAKENAKA_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(11, ws.max_row + 1):
            doc_no = ws[f"E{row}"].value

            if not doc_no or "DETH-NSC" not in str(doc_no):
                continue

            key = base_doc_no(doc_no)

            data[key] = {
                "Takenaka Sheet": sheet,
                "Takenaka Doc No": doc_no,
                "Takenaka Status 1": ws[f"AA{row}"].value,
                "Takenaka Status 2": ws[f"AB{row}"].value,
                "Takenaka Status 3": ws[f"AC{row}"].value,
            }

    return data


def generate_report(tracking_file, takenaka_file):
    takenaka_map = read_takenaka(takenaka_file)

    # NEW: build status summary before openpyxl changes the workbook
    tracking_file.seek(0)
    status_summary_df, status_detail_df = build_status_summary_from_tracking(tracking_file)
    tracking_file.seek(0)

    wb = load_workbook(tracking_file)
    report_sheet = "Open_On_Process_Compare"

    if report_sheet in wb.sheetnames:
        del wb[report_sheet]

    report_ws = wb.create_sheet(report_sheet)

    headers = [
        "Tracking Sheet",
        "Document No",
        "Document Name",
        "Tracking Status",
        "Info",
        "Takenaka Sheet",
        "Takenaka Doc No",
        "Takenaka Status 1",
        "Takenaka Status 2",
        "Takenaka Status 3",
        "Action",
        "Checked Time",
    ]

    report_ws.append(headers)

    for cell in report_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")
    blue = PatternFill(fill_type="solid", fgColor="BDD7EE")

    rows = []
    total_docs = 0
    open_docs = 0

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"B{row}"].value
            doc_name = ws[f"D{row}"].value
            tracking_status = ws[f"F{row}"].value
            info = ws[f"G{row}"].value

            if not doc_no:
                continue

            total_docs += 1

            tracking_status_text = str(tracking_status or "").strip().upper()
            info_text = str(info or "").strip().upper()

            if (
                "OPEN" not in tracking_status_text
                and "ON PROGRESS" not in tracking_status_text
                and "ON PROCESS" not in tracking_status_text
                and "ON PROGRESS" not in info_text
                and "ON PROCESS" not in info_text
            ):
                continue

            open_docs += 1
            key = base_doc_no(doc_no)
            checked_time = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

            if key in takenaka_map:
                src = takenaka_map[key]

                s1 = str(src["Takenaka Status 1"] or "").strip().upper()
                s2 = str(src["Takenaka Status 2"] or "").strip().upper()
                s3 = str(src["Takenaka Status 3"] or "").strip().upper()

                if s1 == "CLOSED":
                    action = "UPDATE TRACKING TO CLOSED"
                    fill = green
                elif s1 == "RETURNED":
                    action = "RETURNED BY NV5 / NEED RESUBMIT"
                    fill = red
                elif s3 == "OVERDUE":
                    action = "OVERDUE / FOLLOW UP"
                    fill = red
                elif s1 == "OPEN" and s2 == "ON PROCESS":
                    action = "OPEN & ON PROCESS"
                    fill = yellow
                elif s1 == "OPEN":
                    action = "OPEN"
                    fill = blue
                else:
                    action = "CHECK"
                    fill = blue

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    src["Takenaka Sheet"],
                    src["Takenaka Doc No"],
                    src["Takenaka Status 1"],
                    src["Takenaka Status 2"],
                    src["Takenaka Status 3"],
                    action,
                    checked_time,
                ]

            else:
                action = "NOT FOUND IN TAKENAKA SOURCE"
                fill = red

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    "",
                    "",
                    "",
                    "",
                    "",
                    action,
                    checked_time,
                ]

            report_ws.append(new_row)
            report_ws[f"K{report_ws.max_row}"].fill = fill

            rows.append(
                {
                    "Tracking Sheet": new_row[0],
                    "Document No": new_row[1],
                    "Document Name": new_row[2],
                    "Tracking Status": new_row[3],
                    "Info": new_row[4],
                    "Takenaka Status 1": new_row[7],
                    "Takenaka Status 2": new_row[8],
                    "Takenaka Status 3": new_row[9],
                    "Action": new_row[10],
                    "Checked Time": new_row[11],
                }
            )

    for col in report_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        report_ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output, total_docs, open_docs, rows, status_summary_df, status_detail_df


# =========================================================
# SESSION
# =========================================================
if "result_df" not in st.session_state:
    st.session_state.result_df = None
    st.session_state.report = None
    st.session_state.total_docs = 0
    st.session_state.open_docs = 0
    st.session_state.action_counts = {}
    st.session_state.last_updated = ""
    st.session_state.status_summary_df = None
    st.session_state.status_detail_df = None

hydrate_session_from_latest()


# =========================================================
# HEADER
# =========================================================
st.markdown(
    f"""
    <div class="hero">
        <div class="hero-grid">
            <div>{render_logo_html()}</div>
            <div>
                <div class="hero-title">Topaz Smart Document Tracker</div>
                <div class="hero-subtitle">TOPAZ BKK1 | ICT Document Control Dashboard</div>
            </div>
            <div class="hero-badge">Shared Dashboard</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# ADMIN UPLOAD
# =========================================================
if st.session_state.role == "admin":
    st.markdown(
        '<div class="info-box">👩‍💼 Admin mode: Admin can upload files, generate dashboard, and update shared data.</div>',
        unsafe_allow_html=True,
    )

    col_upload_1, col_upload_2, col_button = st.columns([1, 1, 0.8])

    with col_upload_1:
        tracking_file = st.file_uploader("1) Upload Tracking_document.xlsx", type=["xlsx"])

    with col_upload_2:
        takenaka_file = st.file_uploader("2) Upload Takenaka Summary.xlsx", type=["xlsx"])

    with col_button:
        st.write("")
        st.write("")
        generate_clicked = st.button("🚀 Generate Dashboard", type="primary", use_container_width=True)

    if tracking_file and takenaka_file and generate_clicked:
        with st.spinner("Reading files and generating dashboard..."):
            report, total_docs, open_docs, rows, status_summary_df, status_detail_df = generate_report(tracking_file, takenaka_file)

        df = pd.DataFrame(rows)
        action_counts = df["Action"].value_counts().to_dict() if not df.empty else {}

        save_latest_dashboard(df, report, total_docs, open_docs, action_counts, status_summary_df, status_detail_df)

        st.session_state.result_df = df
        st.session_state.report = report
        st.session_state.total_docs = total_docs
        st.session_state.open_docs = open_docs
        st.session_state.action_counts = action_counts
        st.session_state.status_summary_df = status_summary_df
        st.session_state.status_detail_df = status_detail_df
        st.session_state.last_updated = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

        st.rerun()

else:
    st.markdown(
        '<div class="viewer-box">ℹ️ Viewer mode: dashboard is read-only. Only Admin can upload files and update shared data.</div>',
        unsafe_allow_html=True,
    )


# =========================================================
# DASHBOARD VIEW
# =========================================================
if st.session_state.result_df is not None:
    df = st.session_state.result_df
    action_counts = st.session_state.action_counts
    last_updated = st.session_state.get("last_updated", "")

    if last_updated:
        st.success(f"Dashboard loaded successfully ✅ | Last updated: {last_updated}")
    else:
        st.success("Dashboard loaded successfully ✅")

    c1, c2, c3, c4, c5 = st.columns(5)

    cards = [
        ("📁", "Total Documents", st.session_state.total_docs, "All documents", "#7c3aed"),
        ("🕒", "Open / On Progress", st.session_state.open_docs, "Documents", "#2563eb"),
        ("▶️", "Open & On Process", action_counts.get("OPEN & ON PROCESS", 0), "Waiting review", "#16a34a"),
        ("🔄", "Need Update", action_counts.get("UPDATE TRACKING TO CLOSED", 0), "Update tracking", "#f97316"),
        ("⚠️", "Overdue", action_counts.get("OVERDUE / FOLLOW UP", 0), "Follow up", "#ef4444"),
    ]

    for col, (icon, title, value, sub, color) in zip([c1, c2, c3, c4, c5], cards):
        with col:
            st.markdown(
                f"""
                <div class="kpi-card" style="--accent:{color};">
                    <div class="kpi-icon">{icon}</div>
                    <div class="kpi-title">{title}</div>
                    <div class="kpi-value">{value}</div>
                    <div class="kpi-sub">{sub}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.write("")
    st.write("")

    chart_col, table_col = st.columns([1.4, 1])

    with chart_col:
        st.markdown('<div class="panel"><div class="panel-title">📊 Action Summary</div>', unsafe_allow_html=True)

        summary_df = pd.DataFrame([{"Action": k, "Count": v} for k, v in action_counts.items()])

        if not summary_df.empty:
            st.bar_chart(summary_df.set_index("Action"))
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
        else:
            st.info("No action data found.")

        st.markdown("</div>", unsafe_allow_html=True)

    with table_col:
        st.markdown('<div class="panel"><div class="panel-title">🧭 Quick Action</div>', unsafe_allow_html=True)

        quick_items = [
            ("Returned by NV5", action_counts.get("RETURNED BY NV5 / NEED RESUBMIT", 0)),
            ("Need update closed", action_counts.get("UPDATE TRACKING TO CLOSED", 0)),
            ("Overdue follow up", action_counts.get("OVERDUE / FOLLOW UP", 0)),
            ("Not found in Takenaka", action_counts.get("NOT FOUND IN TAKENAKA SOURCE", 0)),
        ]

        for label, count in quick_items:
            st.markdown(
                f'<div class="quick-row"><span>{label}</span><span class="count-pill">{count}</span></div>',
                unsafe_allow_html=True,
            )

        st.markdown("</div>", unsafe_allow_html=True)

    # =====================================================
    # V6.1.2 / V6.1.3 / V6.1.4
    # - Color Summary
    # - Export Summary
    # - Drill Down
    # =====================================================
    status_summary_df = st.session_state.get("status_summary_df", None)
    status_detail_df = st.session_state.get("status_detail_df", None)

    st.markdown('<div class="panel"><div class="panel-title">📌 Document Status Summary by Category (Latest Revision)</div>', unsafe_allow_html=True)

    if status_summary_df is not None and not status_summary_df.empty:
        def highlight_status(row):
            status = str(row.get("Status", ""))

            if status == "Open":
                return ["background-color: #dbeafe; color: #1e3a8a; font-weight: 800"] * len(row)
            if status == "On Progress":
                return ["background-color: #fef3c7; color: #92400e; font-weight: 800"] * len(row)
            if status == "Approved":
                return ["background-color: #dcfce7; color: #166534; font-weight: 800"] * len(row)
            if status == "Total Document":
                return ["background-color: #ede9fe; color: #4c1d95; font-weight: 950"] * len(row)

            return [""] * len(row)

        st.dataframe(
            status_summary_df.style.apply(highlight_status, axis=1),
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            label="⬇️ Export Status Summary",
            data=dataframe_to_excel_bytes(status_summary_df, "Status Summary"),
            file_name="Topaz_Status_Summary_By_Category.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

        # Drill Down removed in Clean version.
        # Detailed document list remains available in the Document Action List panel below.

    else:
        st.info("No status summary available. Please generate dashboard again with Tracking_document.xlsx.")

    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="panel"><div class="panel-title">📋 Document Action List</div>', unsafe_allow_html=True)

    filter_col, search_col, download_col = st.columns([1, 2, 0.8])

    with filter_col:
        selected_action = st.selectbox(
            "Filter by Action",
            ["All"] + sorted(df["Action"].dropna().unique().tolist()),
        )

    with search_col:
        search = st.text_input("Search Document No / Document Name")

    filtered_df = df.copy()

    if selected_action != "All":
        filtered_df = filtered_df[filtered_df["Action"] == selected_action]

    if search:
        filtered_df = filtered_df[
            filtered_df.astype(str)
            .apply(lambda x: x.str.contains(search, case=False, na=False))
            .any(axis=1)
        ]

    st.dataframe(filtered_df, use_container_width=True, height=480)

    with download_col:
        st.write("")
        st.write("")

        if st.session_state.report is not None:
            st.download_button(
                label="⬇️ Download Excel Report",
                data=st.session_state.report,
                file_name="Open_On_Process_Compare.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.markdown("</div>", unsafe_allow_html=True)

else:
    st.markdown(
        '<div class="warning-box">⚠️ No shared dashboard data available yet. Please login as Admin and generate dashboard once.</div>',
        unsafe_allow_html=True,
    )
