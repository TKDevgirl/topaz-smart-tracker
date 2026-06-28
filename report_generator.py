from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from config import TAKENAKA_SHEETS, TRACKING_SHEETS
from status_summary import build_status_summary_from_tracking
from utils import base_doc_no


def read_takenaka(takenaka_file):
    wb = load_workbook(takenaka_file, data_only=True)
    data = {}

    for sheet in TAKENAKA_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(11, ws.max_row + 1):
            doc_no = ws[f"E{row}"].value

            if not doc_no or "DETH-NSC" not in str(doc_no):
                continue

            key = base_doc_no(doc_no)

            data[key] = {
                "Takenaka Sheet": sheet,
                "Takenaka Doc No": doc_no,
                "Takenaka Status 1": ws[f"AA{row}"].value,
                "Takenaka Status 2": ws[f"AB{row}"].value,
                "Takenaka Status 3": ws[f"AC{row}"].value,
            }

    return data


def generate_report(tracking_file, takenaka_file):
    takenaka_map = read_takenaka(takenaka_file)

    tracking_file.seek(0)
    status_summary_df, status_detail_df = build_status_summary_from_tracking(tracking_file)
    tracking_file.seek(0)

    wb = load_workbook(tracking_file)
    report_sheet = "Open_On_Process_Compare"

    if report_sheet in wb.sheetnames:
        del wb[report_sheet]

    report_ws = wb.create_sheet(report_sheet)

    headers = [
        "Tracking Sheet",
        "Document No",
        "Document Name",
        "Tracking Status",
        "Info",
        "Takenaka Sheet",
        "Takenaka Doc No",
        "Takenaka Status 1",
        "Takenaka Status 2",
        "Takenaka Status 3",
        "Action",
        "Checked Time",
    ]

    report_ws.append(headers)

    for cell in report_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")
    blue = PatternFill(fill_type="solid", fgColor="BDD7EE")

    rows = []
    total_docs = 0
    open_docs = 0

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"B{row}"].value
            doc_name = ws[f"D{row}"].value
            tracking_status = ws[f"F{row}"].value
            info = ws[f"G{row}"].value

            if not doc_no:
                continue

            total_docs += 1

            tracking_status_text = str(tracking_status or "").strip().upper()
            info_text = str(info or "").strip().upper()

            if (
                "OPEN" not in tracking_status_text
                and "ON PROGRESS" not in tracking_status_text
                and "ON PROCESS" not in tracking_status_text
                and "ON PROGRESS" not in info_text
                and "ON PROCESS" not in info_text
            ):
                continue

            open_docs += 1
            key = base_doc_no(doc_no)
            checked_time = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

            if key in takenaka_map:
                src = takenaka_map[key]

                s1 = str(src["Takenaka Status 1"] or "").strip().upper()
                s2 = str(src["Takenaka Status 2"] or "").strip().upper()
                s3 = str(src["Takenaka Status 3"] or "").strip().upper()

                if s1 == "CLOSED":
                    action = "UPDATE TRACKING TO CLOSED"
                    fill = green
                elif s1 == "RETURNED":
                    action = "RETURNED BY NV5 / NEED RESUBMIT"
                    fill = red
                elif s3 == "OVERDUE":
                    action = "OVERDUE / FOLLOW UP"
                    fill = red
                elif s1 == "OPEN" and s2 == "ON PROCESS":
                    action = "OPEN & ON PROCESS"
                    fill = yellow
                elif s1 == "OPEN":
                    action = "OPEN"
                    fill = blue
                else:
                    action = "CHECK"
                    fill = blue

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    src["Takenaka Sheet"],
                    src["Takenaka Doc No"],
                    src["Takenaka Status 1"],
                    src["Takenaka Status 2"],
                    src["Takenaka Status 3"],
                    action,
                    checked_time,
                ]

            else:
                action = "NOT FOUND IN TAKENAKA SOURCE"
                fill = red

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    "",
                    "",
                    "",
                    "",
                    "",
                    action,
                    checked_time,
                ]

            report_ws.append(new_row)
            report_ws[f"K{report_ws.max_row}"].fill = fill

            rows.append(
                {
                    "Tracking Sheet": new_row[0],
                    "Document No": new_row[1],
                    "Document Name": new_row[2],
                    "Tracking Status": new_row[3],
                    "Info": new_row[4],
                    "Takenaka Status 1": new_row[7],
                    "Takenaka Status 2": new_row[8],
                    "Takenaka Status 3": new_row[9],
                    "Action": new_row[10],
                    "Checked Time": new_row[11],
                }
            )

    for col in report_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        report_ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output, total_docs, open_docs, rows, status_summary_df, status_detail_df
