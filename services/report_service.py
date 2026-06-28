from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from config.columns import TRACKING_COLUMNS
from config.settings import TRACKING_SHEETS
from core.utils import base_doc_no
from services.summary_service import build_status_summary_from_tracking
from services.takenaka_service import read_takenaka_records


def determine_action(src: dict[str, Any]) -> tuple[str, PatternFill]:
    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    yellow = PatternFill(fill_type="solid", fgColor="FFEB9C")
    red = PatternFill(fill_type="solid", fgColor="FFC7CE")
    blue = PatternFill(fill_type="solid", fgColor="BDD7EE")

    s1 = str(src.get("Takenaka Status 1") or "").strip().upper()
    s2 = str(src.get("Takenaka Status 2") or "").strip().upper()
    s3 = str(src.get("Takenaka Status 3") or "").strip().upper()

    if s1 == "CLOSED":
        return "UPDATE TRACKING TO CLOSED", green
    if s1 == "RETURNED":
        return "RETURNED BY NV5 / NEED RESUBMIT", red
    if s3 == "OVERDUE":
        return "OVERDUE / FOLLOW UP", red
    if s1 == "OPEN" and s2 == "ON PROCESS":
        return "OPEN & ON PROCESS", yellow
    if s1 == "OPEN":
        return "OPEN", blue

    return "CHECK", blue


def is_open_or_progress(tracking_status: object, info: object) -> bool:
    tracking_status_text = str(tracking_status or "").strip().upper()
    info_text = str(info or "").strip().upper()

    return (
        "OPEN" in tracking_status_text
        or "ON PROGRESS" in tracking_status_text
        or "ON PROCESS" in tracking_status_text
        or "ON PROGRESS" in info_text
        or "ON PROCESS" in info_text
    )


def generate_report(tracking_file, takenaka_file):
    takenaka_map = read_takenaka_records(takenaka_file)

    tracking_file.seek(0)
    status_summary_df, status_detail_df = build_status_summary_from_tracking(tracking_file)
    tracking_file.seek(0)

    wb = load_workbook(tracking_file)
    report_sheet = "Open_On_Process_Compare"

    if report_sheet in wb.sheetnames:
        del wb[report_sheet]

    report_ws = wb.create_sheet(report_sheet)

    headers = [
        "Tracking Sheet",
        "Document No",
        "Document Name",
        "Tracking Status",
        "Info",
        "Takenaka Sheet",
        "Takenaka Doc No",
        "Takenaka Status 1",
        "Takenaka Status 2",
        "Takenaka Status 3",
        "Action",
        "Checked Time",
    ]

    report_ws.append(headers)

    for cell in report_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    red = PatternFill(fill_type="solid", fgColor="FFC7CE")

    rows: list[dict[str, Any]] = []
    total_docs = 0
    open_docs = 0

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"{TRACKING_COLUMNS['doc_no']}{row}"].value
            doc_name = ws[f"{TRACKING_COLUMNS['doc_name']}{row}"].value
            tracking_status = ws[f"{TRACKING_COLUMNS['status']}{row}"].value
            info = ws[f"{TRACKING_COLUMNS['info']}{row}"].value

            if not doc_no:
                continue

            total_docs += 1

            if not is_open_or_progress(tracking_status, info):
                continue

            open_docs += 1
            key = base_doc_no(doc_no)
            checked_time = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

            if key in takenaka_map:
                src = takenaka_map[key]
                action, fill = determine_action(src)

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    src["Takenaka Sheet"],
                    src["Takenaka Doc No"],
                    src["Takenaka Status 1"],
                    src["Takenaka Status 2"],
                    src["Takenaka Status 3"],
                    action,
                    checked_time,
                ]

            else:
                action = "NOT FOUND IN TAKENAKA SOURCE"
                fill = red

                new_row = [
                    sheet,
                    doc_no,
                    doc_name,
                    tracking_status,
                    info,
                    "",
                    "",
                    "",
                    "",
                    "",
                    action,
                    checked_time,
                ]

            report_ws.append(new_row)
            report_ws[f"K{report_ws.max_row}"].fill = fill

            rows.append(
                {
                    "Tracking Sheet": new_row[0],
                    "Document No": new_row[1],
                    "Document Name": new_row[2],
                    "Tracking Status": new_row[3],
                    "Info": new_row[4],
                    "Takenaka Status 1": new_row[7],
                    "Takenaka Status 2": new_row[8],
                    "Takenaka Status 3": new_row[9],
                    "Action": new_row[10],
                    "Checked Time": new_row[11],
                }
            )

    for col in report_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        report_ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output, total_docs, open_docs, rows, status_summary_df, status_detail_df
