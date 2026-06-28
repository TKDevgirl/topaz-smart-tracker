from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from config.columns import TAKENAKA_COLUMNS
from config.settings import TAKENAKA_SHEETS
from core.utils import base_doc_no


def read_takenaka_records(takenaka_file) -> dict[str, dict[str, Any]]:
    wb = load_workbook(takenaka_file, data_only=True)
    data: dict[str, dict[str, Any]] = {}

    start_row = TAKENAKA_COLUMNS["start_row"]
    doc_no_col = TAKENAKA_COLUMNS["doc_no"]
    status_1_col = TAKENAKA_COLUMNS["status_1"]
    status_2_col = TAKENAKA_COLUMNS["status_2"]
    status_3_col = TAKENAKA_COLUMNS["status_3"]

    for sheet in TAKENAKA_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        for row in range(start_row, ws.max_row + 1):
            doc_no = ws[f"{doc_no_col}{row}"].value

            if not doc_no or "DETH-NSC" not in str(doc_no):
                continue

            key = base_doc_no(doc_no)

            data[key] = {
                "Takenaka Sheet": sheet,
                "Takenaka Doc No": doc_no,
                "Takenaka Status 1": ws[f"{status_1_col}{row}"].value,
                "Takenaka Status 2": ws[f"{status_2_col}{row}"].value,
                "Takenaka Status 3": ws[f"{status_3_col}{row}"].value,
            }

    return data
