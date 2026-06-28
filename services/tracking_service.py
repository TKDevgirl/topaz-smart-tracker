from __future__ import annotations

import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config.columns import STANDARD_CATEGORIES, TRACKING_COLUMNS
from config.settings import TRACKING_SHEETS


def normalize_status(value: Any) -> str:
    text = str(value or "").strip()

    status_map = {
        "open": "Open",
        "on progress": "On Progress",
        "on process": "On Progress",
        "in progress": "On Progress",
        "approved": "Approved",
        "approve": "Approved",
        "close": "Close",
        "closed": "Close",
    }

    return status_map.get(text.lower(), text)


def extract_revision_value(value: Any) -> int:
    text = str(value or "").strip()

    if not text or text.lower() == "nan":
        return 0

    numbers = re.findall(r"\d+", text)
    if numbers:
        return int(numbers[-1])

    if len(text) == 1 and text.isalpha():
        return 1000 + ord(text.upper()) - ord("A") + 1

    return 0


def get_column_by_header_or_fallback(ws, header_name: str, fallback_letter: str) -> str:
    header_name = header_name.strip().lower()

    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == header_name:
            return cell.column_letter

    return fallback_letter


def read_tracking_records(tracking_file) -> pd.DataFrame:
    wb = load_workbook(tracking_file, data_only=True)
    records: list[dict[str, Any]] = []

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        doc_no_col = TRACKING_COLUMNS["doc_no"]
        category_col = get_column_by_header_or_fallback(ws, "Category", TRACKING_COLUMNS["category"])
        doc_name_col = TRACKING_COLUMNS["doc_name"]
        revision_col = get_column_by_header_or_fallback(ws, "version", TRACKING_COLUMNS["revision"])
        status_col = get_column_by_header_or_fallback(ws, "Status", TRACKING_COLUMNS["status"])
        info_col = TRACKING_COLUMNS["info"]

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"{doc_no_col}{row}"].value
            category = ws[f"{category_col}{row}"].value
            doc_name = ws[f"{doc_name_col}{row}"].value
            revision = ws[f"{revision_col}{row}"].value
            status = ws[f"{status_col}{row}"].value
            info = ws[f"{info_col}{row}"].value

            doc_no_text = str(doc_no or "").strip()
            category_text = str(category or "").strip()
            status_text = normalize_status(status)

            if not doc_no_text or doc_no_text.lower() == "nan":
                continue

            if not status_text or status_text.lower() == "nan":
                continue

            if not category_text or category_text.lower() == "nan":
                category_text = sheet

            records.append(
                {
                    "Tracking Sheet": sheet,
                    "Document No": doc_no_text,
                    "Category": category_text,
                    "Document Name": doc_name,
                    "Revision": str(revision or "").strip(),
                    "Revision Sort": extract_revision_value(revision),
                    "Status": status_text,
                    "Info": info,
                    "Excel Row": row,
                }
            )

    return pd.DataFrame(records)


def get_latest_revision_records(records_df: pd.DataFrame) -> pd.DataFrame:
    if records_df.empty:
        return records_df

    return (
        records_df
        .sort_values(["Tracking Sheet", "Document No", "Revision Sort", "Excel Row"])
        .groupby(["Tracking Sheet", "Document No"], as_index=False)
        .tail(1)
        .copy()
    )


def get_standard_categories() -> list[str]:
    return STANDARD_CATEGORIES.copy()
