import re

import pandas as pd
from openpyxl import load_workbook

from config import TRACKING_SHEETS


def normalize_status(value):
    text = str(value or "").strip()

    status_map = {
        "open": "Open",
        "on progress": "On Progress",
        "on process": "On Progress",
        "in progress": "On Progress",
        "approved": "Approved",
        "approve": "Approved",
        "close": "Close",
        "closed": "Close",
    }

    return status_map.get(text.lower(), text)


def get_column_by_name_or_position(ws, preferred_name, fallback_letter):
    preferred_name = preferred_name.strip().lower()

    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == preferred_name:
            return cell.column_letter

    return fallback_letter


def extract_revision_value(value):
    text = str(value or "").strip()

    if not text or text.lower() == "nan":
        return 0

    numbers = re.findall(r"\d+", text)
    if numbers:
        return int(numbers[-1])

    if len(text) == 1 and text.isalpha():
        return 1000 + ord(text.upper()) - ord("A") + 1

    return 0


def build_one_summary(df: pd.DataFrame, sheet_name: str, categories: list[str]) -> pd.DataFrame:
    keep_status = ["Open", "On Progress", "Approved"]
    df = df[df["Status"].isin(keep_status)].copy()

    rows = []

    for status in keep_status:
        status_df = df[df["Status"] == status]
        row = {
            "Document Type": sheet_name,
            "Status": status,
            "Total": len(status_df),
        }

        for cat in categories:
            row[cat] = len(status_df[status_df["Category"] == cat])

        rows.append(row)

    total_row = {
        "Document Type": sheet_name,
        "Status": "Total Document",
        "Total": len(df),
    }

    for cat in categories:
        total_row[cat] = len(df[df["Category"] == cat])

    rows.append(total_row)

    summary_df = pd.DataFrame(rows)

    display_df = summary_df.copy()

    for col in display_df.columns:
        if col not in ["Document Type", "Status"]:
            display_df[col] = display_df[col].apply(lambda x: "-" if int(x) == 0 else int(x))

    return display_df


def build_status_summary_from_tracking(tracking_file):
    """
    Reads RFA/RFI, keeps latest revision per Document No per sheet,
    then builds RFA and RFI category summaries.
    """
    wb = load_workbook(tracking_file, data_only=True)

    records = []

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]

        category_col = get_column_by_name_or_position(ws, "Category", "C")
        status_col = get_column_by_name_or_position(ws, "Status", "F")
        revision_col = get_column_by_name_or_position(ws, "version", "E")

        for row in range(2, ws.max_row + 1):
            doc_no = ws[f"B{row}"].value
            category = ws[f"{category_col}{row}"].value
            doc_name = ws[f"D{row}"].value
            revision = ws[f"{revision_col}{row}"].value
            status = ws[f"{status_col}{row}"].value
            info = ws[f"G{row}"].value

            doc_no_text = str(doc_no or "").strip()
            category_text = str(category or "").strip()
            status_text = normalize_status(status)

            if not doc_no_text or doc_no_text.lower() == "nan":
                continue

            if not status_text or status_text.lower() == "nan":
                continue

            if not category_text or category_text.lower() == "nan":
                category_text = sheet

            records.append(
                {
                    "Tracking Sheet": sheet,
                    "Document No": doc_no_text,
                    "Category": category_text,
                    "Document Name": doc_name,
                    "Revision": str(revision or "").strip(),
                    "Revision Sort": extract_revision_value(revision),
                    "Status": status_text,
                    "Info": info,
                    "Excel Row": row,
                }
            )

    detail_all_df = pd.DataFrame(records)

    if detail_all_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    latest_df = (
        detail_all_df
        .sort_values(["Tracking Sheet", "Document No", "Revision Sort", "Excel Row"])
        .groupby(["Tracking Sheet", "Document No"], as_index=False)
        .tail(1)
        .copy()
    )

    categories = ["MAT", "MCR", "MTS", "CVI", "DWG"]

    rfa_df = latest_df[latest_df["Tracking Sheet"] == "RFA"].copy()
    rfi_df = latest_df[latest_df["Tracking Sheet"] == "RFI"].copy()

    rfa_summary = build_one_summary(rfa_df, "RFA", categories)
    rfi_summary = build_one_summary(rfi_df, "RFI", categories)

    summary_cols = ["Document Type", "Status", "Total", "MAT", "MCR", "MTS", "CVI", "DWG"]
    rfa_summary = rfa_summary[summary_cols]
    rfi_summary = rfi_summary[summary_cols]

    combined_summary = pd.concat([rfa_summary, rfi_summary], ignore_index=True)

    latest_detail_df = latest_df.drop(columns=["Revision Sort", "Excel Row"], errors="ignore")

    return combined_summary, latest_detail_df
